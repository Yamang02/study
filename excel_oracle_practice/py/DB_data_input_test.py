import cx_Oracle
from openpyxl import load_workbook
import os
import ORACLE_DB_connection_test as conn
import logging
import pickle
import uuid
from classes import DataManage

# 스크립트가 위치한 경로
script_path = os.path.dirname(os.path.abspath(__file__))

# EXCEL_TEST.xlsx 파일의 경로
file_path = os.path.join(script_path, "../EXCEL_TEST.xlsx")

# 데이터베이스 연결
connection = conn.connect()

# 로그 설정
logging.basicConfig(filename="app_log.log", level=logging.DEBUG,
                    format="%(asctime)s - %(levelname)s - %(message)s")

# Excel 파일에서 데이터 가져오기
wb = load_workbook(filename=file_path)
ws_sheet1 = wb['Sheet1']
data_list = [{'idx': row[0], 'str': row[1]} for row in ws_sheet1.iter_rows(min_row=2, values_only=True)]


# create_data_list 함수를 사용하여 데이터 리스트 생성
data_list = DataManage.create_data_list(ws_sheet1) 

# 유니크한 파일 이름 생성
unique_filename = str(uuid.uuid4()) + ".pkl"

# 데이터 리스트를 피클 파일에 저장
with open(unique_filename, "wb") as f:
    pickle.dump(data_list, f)



# 데이터 삽입 쿼리 실행
try:
    cursor = connection.cursor()

    for data in data_list:
        query = "INSERT INTO EXCEL_TEST (PK_TYPE_NUM, TYPE_VARCHAR) VALUES (:1, :2)"
        cursor.execute(query, (data['idx'], data['str']))
        logging.info(f"데이터 삽입 성공: {data}")

    connection.commit()
    print("데이터가 성공적으로 삽입되었습니다.")
except cx_Oracle.Error as error:
    print(f"데이터 삽입 중 오류 발생: {error}")
    connection.rollback()
    logging.error(f"데이터 삽입 중 오류 발생: {error}")
finally:
    cursor.close()
    connection.close()
