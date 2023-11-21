import openpyxl

class ExcelDataManager :

    #컬럼 key 값들을 가져오는 메소드
    def get_excel_key_value(self, workbook, sheet_name):
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            return [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)]
        else:
            return []

     #해당 시트의 헤더 리스트를 배열로 반환
    def worksheetHeader(self, worksheet):
        header_data = []

        for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            header_data = list(row)

        return header_data

    #객체 인스턴스를 특정 시트에 추가하는 메소드
    def add_instance_to_excel_sheet(self, workbook, sheet_name, instance, row_number=None):
        # print(instance)
        if not hasattr(instance, '__dict__'):
            raise ValueError("Input must be an object instance.")

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            header_row = self.worksheetHeader(sheet)
        else:
            print("creating_new_sheet...")
            sheet = workbook.create_sheet(title=sheet_name)
            header_row = []

            for field_name, field_value in instance.__dict__.items():
                header_row.append(field_name)
                sheet.cell(row=1, column=len(header_row), value=field_name)

        if row_number is None:
            row_number = sheet.max_row + 1

        data_row = []
        # print(header_row)

        for field_name in header_row:
            # print(field_name)
            field_value = getattr(instance, field_name)
            if isinstance(field_value, (int, float)):
                data_row.append(str(field_value))  # 숫자 값을 문자열로 변환하여 추가
            else:
                data_row.append(field_value)

            sheet.cell(row=row_number, column=header_row.index(field_name) + 1, value=str(field_value))


        return sheet, row_number



    # 새 정보 입력
    def dataInsert(self, worksheet, target_row : int, data):
        data_dict = data.__dict__

        for col_idx, (key, value) in enumerate(data_dict.items(), 1):
            worksheet.cell(row=target_row, column=col_idx).value = value

    # 특정 행의 정보 확인
    def dataSelect(self, worksheet, target_row : int):
        selected_row_data = {}
        header_row = worksheet[1]  # 1행의 데이터를 가져옴

        for col_idx, cell in enumerate(worksheet.iter_rows(min_row=target_row, max_row=target_row, values_only=True), 1):
            for header_cell, value in zip(header_row, cell):
                selected_row_data[header_cell.value] = value
                print(f"{header_cell.value}, value: {value}")

        return selected_row_data
    

     # 특정 행의 정보 수정 (컬럼 대소문자 구분)

    # 데이터 업데이트 
    def dataUpdate(self, worksheet, target_row : int, data):
        data_dict = data.__dict__
        header_row = worksheet[1]  # 1행의 데이터를 가져옴

        for col_idx, (key, value) in enumerate(data_dict.items(), 1):
            # 헤더에 해당하는 컬럼을 찾기 위해 key와 일치하는 컬럼을 탐색
            for header_cell in header_row:
                if header_cell.value == key:
                    worksheet.cell(row=target_row, column=header_cell.column, value=value)
                    break
            else:
                print(f"Warning: '{key}' is not a valid column in the worksheet.")

     # 행의 'visible' 값을 'n'으로 변환하고 변경된 행의 인덱스를 반환하는 함수

    # visible값 토글
    def visibleToggle(self, worksheet, target_row : int):
        header_row = worksheet[1]  # 1행의 데이터를 가져옴

        # 'visible' 컬럼을 찾기 위해 헤더를 탐색
        visible_column_idx = None
        for col_idx, header_cell in enumerate(header_row, 1):
            if header_cell.value == 'visible':
                visible_column_idx = col_idx
                break

        if visible_column_idx is None:
            print("Error: 'visible' column not found in the worksheet.")
            return None

        # 현재 'visible' 값을 확인하여 토글
        current_value = worksheet.cell(row=target_row, column=visible_column_idx).value
        new_value = 'n' if current_value == 'y' else 'y'
        worksheet.cell(row=target_row, column=visible_column_idx, value=new_value)
        return target_row
    
         # visible 값이 'n'인 행을 삭제하고 삭제된 행의 개수를 반환하는 함수
    
    # 데이터 삭제
    def dataDelete(self, worksheet):
        header_row = worksheet[1]  # 1행의 데이터를 가져옴

        # 'visible' 컬럼을 찾기 위해 헤더를 탐색
        visible_column_idx = None
        for col_idx, header_cell in enumerate(header_row, 1):
            if header_cell.value == 'visible':
                visible_column_idx = col_idx
                break

        if visible_column_idx is None:
            print("Error: 'visible' column not found in the worksheet.")
            return None

        # visible 값이 'n'인 행을 찾아 삭제
        deleted_row_count = 0
        for row_idx, cell in enumerate(worksheet.iter_rows(min_row=2, min_col=visible_column_idx,
                                                           max_col=visible_column_idx, values_only=True), 2):
            if cell[0] == 'n':
                worksheet.delete_rows(row_idx)
                deleted_row_count += 1

        return deleted_row_count
    
    def create_data_list(self, worksheet) -> list:
        headers = [cell for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        data_list = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data = {headers[i]: row[i] for i in range(len(headers))}
            data_list.append(data)
        return data_list

class data_type_TEST :
    def __init__(self, num, name) :
        self.num = num
        self.name = name
        self.visible = 'y'
